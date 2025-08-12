from extras.scripts import *

from virtualization.models import VirtualMachine
from virtualization.choices import VirtualMachineStatusChoices
from tenancy.models import Tenant
from extras.models import CustomFieldChoiceSet

from openpyxl import Workbook
from openpyxl.styles import Font, colors
from openpyxl.worksheet.table import Table, TableStyleInfo

import datetime
import os

# --------------------

# file
filename = "NetboxOut_" + str(datetime.datetime.now().strftime("%Y-%m")) + '.xlsx'
savePath = os.path.join('/opt/netbox/netbox/media',filename)

# prices
priceCpu = 10.45
priceRam = 4.14
priceStrg = 0.21

# ___VM attributes___
# Choices are: _name, bookmarks, cluster, cluster_id, comments, config_template, config_template_id, contacts, created, \
#  custom_field_data, description, device, device_id, disk, id, images, interface_count, interfaces, journal_entries, \
#  last_updated, local_context_data, memory, name, platform, platform_id, primary_ip4, primary_ip4_id, primary_ip6, \
#  primary_ip6_id, role, role_id, serial, services, site, site_id, status, subscriptions, tagged_items, tags, tenant, \
#  tenant_id, vcpus, virtual_disk_count, virtualdisks

# --------------------

class Application:
     def __init__(self, tenant, name, cost, site, vm, cluster, role, platform, desc, cores, ram, storage):
         self.tenant = tenant
         self.name = name # = app
         self.cost = cost
         self.site = site
         self.vm = vm
         self.cluster = cluster
         self.role = role
         self.platform = platform
         self.desc = desc
         self.cores = cores
         self.ram = ram
         self.storage = storage
     def get_tenant(self):
         return self.tenant
     def get_name(self):
         return self.name
     def get_cost(self):
         return self.cost
     def get_site(self):
         return self.site
     def get_vm(self):
         return self.vm
     def get_cluster(self):
         return self.cluster
     def get_role(self):
         return self.role
     def get_platform(self):
         return self.platform
     def get_desc(self):
         return self.desc
     def get_cores(self):
         return self.cores
     def get_ram(self):
         return self.ram
     def get_storage(self):
         return self.storage

class TenantCalc:
    def __init__(self, id, name):
        self.id = id
        self.name = name
    def get_id(self):
        return self.id
    def get_name(self):
        return self.name

class ExportAllVMResourcesToXLSX(Script):
    class Meta:
        name = "Export All Tenant Resources to .xlsx"
        description = "Check assigned hardware resources of all active VMs and export them to an Excel file sorted by tenant."
        commit_default = False

    # --------------------

    @staticmethod
    def last_row_to_function(workbook, maxrow):
        ws = workbook
        lastRow = f"A{ws.max_row}:{maxrow}{ws.max_row}"
        for row in ws[lastRow]:
            for cell in row:
                cell.font = Font(bold=True)
                if str(cell.value).startswith('='):
                    cell.data_type = 'f'

    def run(self, data, commit):

        # tenants to list
        tenants = []
        for tenant in Tenant.objects.all():
            tenants.append(TenantCalc(tenant.id, tenant.name))
        #tenants = tenants.reverse() # reverse order -> NOPE, breaks typing ^^
        tenants = list(reversed(tenants))
        if len(tenants) == 0:
            self.log_failure(f"No tenants / mandates found! Script will be terminated.")
            quit()
        self.log_info(f"Tenants collected.")

        # app display name lookup table
        appLookup = []
        for choiceSet in CustomFieldChoiceSet.objects.filter(name='application'):
            for choice in choiceSet.choices:
                appLookup.append(choice)

        # iterate through active VMs (and add resources to tenant + application pairs -> Nope)
        applications = []
        for vm in VirtualMachine.objects.filter(status=VirtualMachineStatusChoices.STATUS_ACTIVE):
            customData = vm.get_custom_fields()
            splits = ", ".join("=".join((str(k), str(v))) for k, v in customData.items())
            app = splits.split('Application=')[1].split(',')[0]
            cost = splits.split('Cost center=')[1].split(',')[0]

            # app -> Label
            for entry in appLookup:
                if app == entry[0]:
                    app = entry[1]

            if cost != 'None':
                cost = int(cost)
            for tenant in tenants:
                if vm.tenant_id == tenant.get_id():
                    # value handling
                    if vm.vcpus == 0.00 or vm.vcpus is None:
                        vcpusX = 0.00
                    else:
                        vcpusX = vm.vcpus
                    if vm.memory == 0 or vm.memory is None:
                        memoryX = 0
                    else:
                        if vm.memory % 1024 != 0:
                            memoryX = round(vm.memory/1000)
                        else:
                            memoryX = round(vm.memory/1024)
                        # memoryX = vm.memory
                    if vm.disk == 0 or vm.disk is None:
                        diskX = 0
                    else:
                        diskX = round(vm.disk/1000)
                    if str(vm.role) == "" or vm.site is None:
                        roleX = "None"
                    else:
                        roleX = str(vm.role)
                    if str(vm.platform) == "" or vm.platform is None:
                        platformX = "None"
                    else:
                        platformX = str(vm.platform)
                    if vm.description == "" or vm.description is None:
                        descX = "None"
                    else:
                        descX = vm.description
                    applications.append(Application(tenant.get_name(), app, cost, str(vm.site), vm.name, str(vm.cluster),
                                                    roleX, platformX, descX, vcpusX, memoryX, diskX))
        self.log_info(f"Resources collected.")

        # setup Workbook for Excel output & add data
        wb = Workbook()
        ws = wb.active
        headRow = ["Tenant", "Application", "Cost Center", "Site", "Virtual Machine", "Cluster", "Role", "Platform",
                   "vCores (per core)", "RAM (per GB)", "Storage (per GB)", "Description"]
        ws.append(headRow)
        for app in applications:
            ws.append([app.get_tenant(),app.get_name(),app.get_cost(),app.get_site(),app.get_vm(),app.get_cluster(),
                       app.get_role(), app.get_platform(),app.get_cores(),app.get_ram(),app.get_storage(),app.get_desc()])

        # last row + styling & mark functions as such to prevent errors
        emptyRow = ["", "", "", "", "", "", "", "", "", "", "", ""] # dumb but it looks better (:
        ws.append(emptyRow)
        bottomRow = ["", "", "", "", "", "", "", "Gesamtanzahl:",
                     f"=SUBTOTAL(9,Resources[vCores (per core)])",
                     f"=SUBTOTAL(9,Resources[RAM (per GB)])",
                     f"=SUBTOTAL(9,Resources[Storage (per GB)])", ""]
        ws.append(bottomRow)
        # lastRow = f"A{ws.max_row}:L{ws.max_row}" # ---> AS FUNCTION CALL
        # for row in ws[lastRow]:
        #     for cell in row:
        #         cell.font = Font(bold=True)
        #         if str(cell.value).startswith('='):
        #             cell.data_type = 'f'

        # additional rows for subtotal calculations with pricing
        bottomRow2 = ["", "", "", "", "", "", "", "Einzelpreis:",
                     priceCpu,priceRam,priceStrg,""]
        ws.append(bottomRow2)
        bottomRow3 = ["", "", "", "", "", "", "", "Teilsumme:",
                      f"=OFFSET(INDIRECT(ADDRESS(ROW(), COLUMN())), -1, 0) * OFFSET(INDIRECT(ADDRESS(ROW(), COLUMN())), -2, 0)",
                      f"==OFFSET(INDIRECT(ADDRESS(ROW(), COLUMN())), -1, 0) * OFFSET(INDIRECT(ADDRESS(ROW(), COLUMN())), -2, 0)",
                      f"==OFFSET(INDIRECT(ADDRESS(ROW(), COLUMN())), -1, 0) * OFFSET(INDIRECT(ADDRESS(ROW(), COLUMN())), -2, 0)",
                      ""]
        ws.append(bottomRow3)
        self.last_row_to_function(ws,'L')
        bottomRow4 = ["", "", "", "", "", "", "", "Gesamtsumme:",
                      f"=OFFSET(INDIRECT(ADDRESS(ROW(), COLUMN())), -1, 0) + OFFSET(INDIRECT(ADDRESS(ROW(), COLUMN())), -1, 1) + OFFSET(INDIRECT(ADDRESS(ROW(), COLUMN())), -1, 2)" ,
                      "", "", ""]
        ws.append(bottomRow4)
        self.last_row_to_function(ws, 'L')

        # change column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if len(str(cell.value)) > max_length and not str(cell.value).startswith('='):
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.05
            ws.column_dimensions[column_letter].width = adjusted_width

        # head row style & sheet name
        ft = Font(name='Calibri', size=12, bold=True, color='ffece9e4') # color = argb hex value
        headRowx = "A1:L1"
        for row in ws[headRowx]:
            for cell in row:
                cell.font = ft
        sheetName = f"Tenant Resources {datetime.datetime.now().strftime("%d.%m.%Y")}"
        ws.title = sheetName

        # [UPDATE] format as table // https://openpyxl.readthedocs.io/en/3.1.3/worksheet_tables.html
        tab = Table(displayName="Resources", ref=f"A1:L{len(applications)+1}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # save to file
        self.log_info(f"Saving file: {savePath}")
        wb.close()
        wb.save(savePath)
        self.log_success(f"File exported successfully.")

        # export to SFTP ?
        # https://stackoverflow.com/questions/33751854/upload-file-via-sftp-with-python#73432631
        # import paramiko
        #
        # with paramiko.SSHClient() as ssh:
        #     ssh.load_system_host_keys()
        #     ssh.connect(host, username=username, password=password)
        #     sftp = ssh.open_sftp()
        #     sftp.chdir('public')
        #     sftp.put('C:\Users\XXX\Dropbox\test.txt', 'test.txt')


